from Functions import *
import openpyxl

def h1(user):
    """
    First Hash function.
 
    Return a value based in Ascii above the first char in 'user'

    Parameters:
    user -- The username (should be different of 'ñ' and 'Ñ')

    Exceptions:
    ValueError -- user[0] is not a letter different of 'ñ' and 'Ñ'
   
    """
    if (user[0] not in "qwertyuiopasdfghjklzxcvbnmQWERTYUIOPASDFGHJKLZXCVBNM"):
        raise ValueError("the first char is not a letter different of 'ñ' and 'Ñ'")
    return 1000*(ord(user[0])-65)+1

# Second Hash function
def h2(usuario):
    return len(usuario)-4


def NewUserAndPassword():
    # Pedir datos
    console_print("Please register")
    # Select the user
    user = input("User: ")
    password = input("Password: ")
    # Select a valid user and password
    i = 0
    while(user[0] not in "qwertyuiopasdfghjklzxcvbnmQWERTYUIOPASDFGHJKLZXCVBNM" or len(user)<5 or len(password)<5):
        # Short username
        if (len(user)<5):
            console_print("Please use an user with 5 character or more")
        # Short password
        if (len(password)<5):
            console_print("Please use a stronger password with 5 character or more")
        # Error with Hash System
        if(user[0] not in "qwertyuiopasdfghjklzxcvbnmQWERTYUIOPASDFGHJKLZXCVBNM"):
            console_print("Warning: You should use a letter minus 'ñ' or 'Ñ'")
        i+=1
        # Too many tries
        if(i > 3):
            return False
        user = input("User: ")
        password = input("Password: ")
    return (user, password)

def register():
    # Abrir documento
    documento = openpyxl.load_workbook('StrangeMenuData.xlsx')
    UserData = documento["User"]
    PasswordData = documento["Password"]
   
    # Select a new valid user and password
    Data = NewUserAndPassword() 
    if Data == False:
        return console_print("Too many tries, try later.")
    user = Data[0]
    password = Data[1]

    # Hash System and data comprobation
    i = h1(user)
    j = h2(user)
    while(UserData.cell(i,j).value != None): # Checking the place
        if (UserData.cell(i,j).value == user): # Do not have repeats usernames
            console_print("This username is already taken.")
            Data = NewUserAndPassword()
            if Data == False: # Error selecting the user and password
                return console_print("Too many tries, try later")
            user = Data[0]
            password = Data[1]
            i = h1(user)
            j = h2(user)
        else:
            i+=1 # Lineal probing

    # Successful registration
    UserData.cell(i,j).value = user
    PasswordData.cell(i,j).value = password
    console_print("You are registered now")
    documento.save('StrangeMenuData.xlsx')
    return

def login():
    # Abrir documento
    documento = openpyxl.load_workbook('StrangeMenuData.xlsx')
    UserData = documento["User"]
    PasswordData = documento["Password"]

    # Pedir datos
    for k in range(3):
        console_print("Please login")
        user = input("User: ")
        password = input("Password: ")
        i = h1(user)
        j = h2(user)
        # Search the username in data
        while (UserData.cell(i,j).value != None and UserData.cell(i,j).value != user):
            i+=1
        if(UserData.cell(i,j).value == None): # User not founded
            console_print(f"User not founded, you have {2-k} more tries")
        elif(UserData.cell(i,j).value == user): # User founded
            if(PasswordData.cell(i,j).value != password): # the password does not match
                console_print(f"Username and password do not match, you have {2-k} more tries")
            else: # User and password match
                console_print("Successful login")
                documento.save('StrangeMenuData.xlsx') # make sure to close the document
                return (user, password)
    documento.save('StrangeMenuData.xlsx') # make sure to close the document
    return False

def MenuForo(user):
    # Open the document
    documento = openpyxl.load_workbook('StrangeMenuData.xlsx')
    ForumData = documento["Forum"]

    # Deploy last 10 chat
    j = ForumData.cell(1,3).value
    console_print("10 recent messages")
    for i in range(0,10):
        k = (j + i - 1) % 10 + 1
        print(f"[{ForumData.cell(k,4).value}] {ForumData.cell(k,1).value}: {ForumData.cell(k,2).value}")
    time.sleep(1)
    print("\t1) Write a new message")
    print("\t2) Exit")
    aux = getkey()
    if aux == 1:
        # Select the new message
        message = input("Write a message: ")

        # Change the actual line and actualizate the auxiliar counter
        j = ForumData.cell(1,3).value
        ForumData.cell(j,1).value = user
        ForumData.cell(j,2).value = message
        ForumData.cell(j,4).value = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        ForumData.cell(1,3).value = j % 10 + 1
        console_print("Message writed")
      
    documento.save('StrangeMenuData.xlsx') # make sure to close the document
    return

def ChangePassword(user, password):
    posible_password = input("Insert your current password: ")
    if posible_password != password:
        return console_print("ERROR: The password do not match, try later")
    else:
        # Open the document
        documento = openpyxl.load_workbook('StrangeMenuData.xlsx')
        UserData = documento["User"]
        PasswordData = documento["Password"]
        i = h1(user)
        j = h2(user)
        # Search the username in data
        while (UserData.cell(i,j).value != user):
            i+=1
        # Updating the password
        PasswordData.cell(i,j).value = input("Insert your new password: ")

        documento.save('StrangeMenuData.xlsx') # make sure to close the document
        return console_print("You have changed your password")

def WriteDescription(user):
    # Open the document
    documento = openpyxl.load_workbook('StrangeMenuData.xlsx')
    UserData = documento["User"]
    DescriptionData = documento["Description"]
    i = h1(user)
    j = h2(user)
    # Search the username in data
    while (UserData.cell(i,j).value != user):
        i+=1
    print("Your current description is:", DescriptionData.cell(i,j).value)
    if(input("Do you want to actualizate it ? (yes/no): ") == "yes"):
        DescriptionData.cell(i,j).value = input("Write your new description: ")
        console_print("You have actualizated your description")
    documento.save('StrangeMenuData.xlsx') # make sure to close the document
    return      
   
def MenuDiegoSapp(user, password):
    Flag = True
    while(Flag):
        console_print("Welcome to the DiegoSapp Menu!")
        print("\t1) Join to the forum")
        print("\t2) Change the description")
        print("\t3) Change the password")
        print("\t4) Exit")
        aux = getkey()
        if aux == 1:
            MenuForo(user)
        if aux == 2:
            WriteDescription(user)
        if aux == 3:
            ChangePassword(user, password)
        if aux not in range(1,4):
            Flag = False
    return

def DiegoSapp():
    flag = True
    while(flag): 
        console_print("Welcome to DiegoSapp!")
        print("\t1) Register.")
        print("\t2) Login.")
        print("\t3) Exit.")
        aux = getkey()
        if aux not in range(1,3): # To Exit
            flag = False
        if aux == 1: # To register
            register() # Create a new user-password or warning if is already created
        if aux == 2: # To login
            Session = login() # Login with a created user-password
            if (Session == False):
                console_print("Too many errors with the login")
                flag = False
            else:
                MenuDiegoSapp(Session[0], Session[1])
                flag = False
    return console_print("Nice to see you in DiegoSapp")